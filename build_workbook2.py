#!/usr/bin/env python3
"""
build_workbook2.py — organize all results into one clean workbook modeled on the
mentor's ResultsData.xlsx (versioned experiment log), but with both R^2 and RMSE,
plain wording (no all-caps emphasis, no em dashes), and extra sheets for the
deployables, confidence, per-metal and per-pair analysis, the Zhang comparison,
and methods. Saves REE_Results_Organized.xlsx.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEAD = PatternFill("solid", fgColor="1F4E79"); SUB = PatternFill("solid", fgColor="D5E8F0")
WHITE = Font(color="FFFFFF", bold=True, name="Arial", sz=11); BOLD = Font(bold=True, name="Arial", sz=11)
REG = Font(name="Arial", sz=10); THIN = Side(style="thin", color="BFBFBF")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN); WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")

wb = Workbook()

def sheet(title, headers, rows, widths, note=None, freeze="A2"):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    if ws.title == "Sheet": ws.title = title
    r0 = 1
    if note:
        ws.cell(1, 1, note).font = Font(italic=True, name="Arial", sz=9, color="555555")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers)); r0 = 2; freeze = "A3"
    for j, h in enumerate(headers, 1):
        c = ws.cell(r0, j, h); c.fill = HEAD; c.font = WHITE; c.alignment = WRAP; c.border = BORD
    for i, row in enumerate(rows, r0 + 1):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v); c.font = REG; c.border = BORD; c.alignment = WRAP
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = freeze
    return ws

# 1. Summary of deployable models
sheet("Summary",
    ["Model (deployable)", "Use case", "Split (honest)", "R2", "RMSE", "Top 25% R2", "Top 25% RMSE", "Top 10% R2", "Top 10% RMSE", "Conformal 90% coverage"],
    [["Track A: single LightGBM", "Screen a new extractant before testing it", "Molecule-grouped (new molecule)", 0.466, 1.148, 0.825, 0.642, 0.912, 0.493, 0.89],
     ["Track B: NNLS stack of trees", "Optimize conditions for a known extractant", "Random row (new conditions)", 0.725, 0.823, 0.911, 0.424, 0.940, 0.341, 0.90]],
    [26, 38, 30, 8, 8, 11, 12, 11, 12, 16],
    note="Both metrics shown throughout: R2 (higher is better) and RMSE in log units (lower is better). Data cleaned to 7065 rows (dropped exact duplicates and replicate groups disagreeing by more than 2 log units).")

# 2. Experiment log (mentor template style, our modeling work)
sheet("Experiment Log",
    ["Version", "Extractant representation", "Metals", "Solvents", "Acid", "Model", "Settings", "CV R2", "CV RMSE", "Notes"],
    [
    ["A1", "Conditions only (no structure)", "Properties", "Properties", "Properties", "LightGBM", "new-molecule grouped CV", 0.466, 1.148, "Track A baseline. Structure does not generalize from ~300 molecules, so conditions plus metal carry it."],
    ["A2", "Conditions only (no structure)", "Properties", "Properties", "Properties", "NNLS stack (lgb, xgb, hgb, et, cat, ridge)", "new-molecule grouped CV", 0.504, 1.106, "Stack adds accuracy but makes errors less rankable, so confidence drops. Single model chosen for Track A."],
    ["B1", "Conditions + ECFP + ligand", "Properties", "Properties", "Properties", "LightGBM", "known-molecule random CV", 0.714, 0.831, "Track B single model."],
    ["B2", "Conditions + ECFP + ligand", "Properties", "Properties", "Properties", "NNLS stack (lgb 0.40, et 0.43, xgb 0.14, cat 0.03)", "known-molecule random CV", 0.725, 0.823, "Track B deployable. Stack wins accuracy and confidence here. Beats the mentor log best test RMSE of about 0.96."],
    ["C1", "Conditions + ECFP + RDKit descriptors", "Properties", "Properties", "Properties", "LightGBM", "known-molecule random CV", 0.70, 0.86, "RDKit descriptors did not help over ECFP and slightly hurt."],
    ["C2", "Conditions + ECFP + learned embeddings", "Properties", "Properties", "Properties", "LightGBM", "known-molecule random CV", 0.69, 0.88, "Pretrained embeddings hurt. ECFP plus conditions is the better representation."],
    ["G1", "Graph (Chemprop D-MPNN) alone", "Properties", "Properties", "Properties", "Message-passing GNN", "new-molecule grouped CV", 0.23, None, "GNN alone is weak on ~300 molecules. RMSE not logged."],
    ["G2", "Graph GNN + tree global blend", "Properties", "Properties", "Properties", "GNN + LightGBM blend", "new-molecule grouped CV", 0.47, 1.15, "GNN adds no generalizable gain over the tree model."],
    ["T1", "Conditions + ECFP (PCA 50) + ligand", "Properties", "Properties", "Properties", "TabPFN global (context capped 1000)", "known-molecule random CV", 0.36, None, "TabPFN cannot use more than 1000 context rows on CPU, so a single global TabPFN is weak here."],
    ["T2", "Conditions + ECFP (PCA 50) + ligand", "Properties", "Properties", "Properties", "TabPFN local experts + tree stack", "known-molecule random CV", 0.701, None, "Lighter in-stack test (6 regions, 3 folds). TabPFN earns NNLS weight 0.13 but adds only +0.0011 R2 over the trees-only stack (0.700); error-correlation with the trees is 0.82, so little diversity. Dropped, since it does not clear the keep threshold."],
    ["P1", "Conditions + ECFP + slope or physics features", "Properties", "Properties", "Properties", "LightGBM", "known-molecule random CV", 0.72, 0.83, "Physics-informed features gave no clear gain. The clean logK + n log[HA] + n pH signal is not present in most rows."],
    ["H1", "Conditions + ECFP + ligand", "Properties", "Properties", "Properties", "LightGBM tuned (randomized search)", "both tracks", 0.484, 1.13, "Hyperparameter search moved Track A 0.466 to 0.484 and Track B not at all. Negligible."],
    ["E1", "Conditions + ECFP + ligand, tail experts", "Properties", "Properties", "Properties", "LightGBM + extreme-region experts", "known-molecule random CV", 0.70, 0.86, "Tail experts for the extremes hurt overall (0.725 to about 0.70). The global model compresses tails but naive experts do not fix it."],
    ["X0", "Leaked single split (for contrast only)", "Properties", "Properties", "Properties", "LightGBM", "single split with molecule leakage", 0.60, 1.05, "Not honest. A molecule appears in train and test. Shown only to explain why earlier numbers looked higher."],
    ],
    [10, 34, 12, 12, 10, 34, 22, 8, 9, 46],
    note="Versioned log in the style of ResultsData.xlsx, with both CV R2 and CV RMSE. CV numbers are cross-validated (honest), not a single lucky split.")

# 3. Model comparison: single vs stack, members and weights
sheet("Model Comparison",
    ["Track", "Member", "Member R2", "NNLS weight", "Single R2", "Single RMSE", "Stack R2", "Stack RMSE", "Chosen for deploy"],
    [
    ["A (new molecule)", "cat", 0.491, 0.37, 0.466, 1.148, 0.504, 1.106, "Single (better confidence)"],
    ["A (new molecule)", "et", 0.470, 0.24, "", "", "", "", ""],
    ["A (new molecule)", "xgb", 0.466, 0.16, "", "", "", "", ""],
    ["A (new molecule)", "lgb", 0.459, 0.11, "", "", "", "", ""],
    ["A (new molecule)", "ridge", 0.178, 0.10, "", "", "", "", ""],
    ["A (new molecule)", "hgb / mlp", 0.47, 0.02, "", "", "", "", ""],
    ["B (known molecule)", "et", 0.708, 0.43, 0.714, 0.831, 0.725, 0.823, "Stack (better accuracy and confidence)"],
    ["B (known molecule)", "lgb", 0.711, 0.40, "", "", "", "", ""],
    ["B (known molecule)", "xgb", 0.707, 0.14, "", "", "", "", ""],
    ["B (known molecule)", "cat", 0.692, 0.03, "", "", "", "", ""],
    ["B (known molecule)", "hgb / mlp / ridge", 0.68, 0.00, "", "", "", "", "Dropped by NNLS"],
    ],
    [18, 16, 11, 12, 9, 10, 9, 10, 32],
    note="NNLS stacking gives zero weight to members that do not improve the blend, so weak models drop out automatically.")

# 4. Confidence bakeoff
sheet("Confidence",
    ["Track", "Predictor", "Err-model recipe", "Spearman(err, abs error)", "Top 50% R2", "Top 50% RMSE", "Top 25% R2", "Top 25% RMSE", "Top 10% R2", "Top 10% RMSE", "Chosen"],
    [
    ["A", "single", "plain + strong", 0.416, 0.719, 0.812, 0.825, 0.642, 0.912, 0.493, "yes"],
    ["A", "stack", "plain + strong", 0.380, 0.718, 0.791, 0.802, 0.656, 0.843, 0.549, ""],
    ["A", "stack", "plain + regularized", 0.387, 0.727, 0.776, 0.812, 0.634, 0.844, 0.545, ""],
    ["A", "stack", "rich (members, disagreement, novelty)", 0.338, 0.687, 0.835, 0.799, 0.662, 0.863, 0.569, "worst"],
    ["A", "stack", "lean (disagreement, novelty)", 0.340, 0.713, 0.804, 0.790, 0.670, 0.831, 0.556, ""],
    ["B", "single", "plain + strong", 0.426, 0.866, 0.550, 0.915, 0.440, 0.938, 0.379, ""],
    ["B", "stack", "plain + regularized", 0.449, 0.871, 0.535, 0.911, 0.424, 0.940, 0.341, "yes"],
    ["B", "stack", "rich", 0.429, 0.872, 0.542, 0.923, 0.408, 0.937, 0.339, ""],
    ["B", "stack", "lean", 0.437, 0.875, 0.535, 0.908, 0.440, 0.942, 0.344, ""],
    ],
    [7, 9, 34, 18, 11, 12, 11, 12, 11, 12, 9],
    note="Confidence is a learned error model (err_lgb) that predicts each row's absolute error. Rows are ranked by it. Rich features (member predictions, disagreement, novelty) hurt, so the plain recipe was kept. Conformal intervals scale by this error and hit their target coverage.")

# 5 and 6: metal and pair from CSV
bm = pd.read_csv("metal_confidence_by_metal.csv")
sheet("By Metal",
    ["Metal", "n", "R2", "RMSE", "Median confidence error", "Top 25% R2", "Top 25% RMSE"],
    bm.values.tolist(), [10, 7, 8, 8, 20, 11, 12],
    note="Track B (known molecule). Accuracy and confidence by metal, both metrics. The confidence error tracks RMSE, so the model flags the metals it predicts poorly (for example Np(V), U(VI), Er(III)).")
bp = pd.read_csv("metal_confidence_by_pair.csv")
sheet("By Metal Pair",
    ["Metal pair", "n", "Separation R2", "Separation RMSE", "Mean confidence error"],
    bp.values.tolist(), [18, 7, 14, 16, 20],
    note="Predicted logD difference between two metals at the same conditions (the separation use case). Overall separation R2 0.599, RMSE 1.025. Tight adjacent pairs are hardest and the confidence error is highest there.")

# 7. Zhang comparison
sheet("Zhang Comparison",
    ["Aspect", "Dr. Zhang (SAFE-MolGen)", "Our model"],
    [
    ["Goal", "Generate new extractants with an LLM, then screen them", "Predict logD, optimize conditions, and predict separations"],
    ["Supervised model", "XGBoost classifier (also NN with CV, random forest)", "LightGBM and an NNLS tree stack, plus matching classifiers"],
    ["Task type", "3-class classification only", "Regression of continuous logD, plus 3-class and binary classifiers"],
    ["Data", "8075 rows, f-element extraction (ACSEPT, DGA, IDEaL, ORNL)", "Same 8075 rows (7065 after cleaning)"],
    ["Features", "1860 (Morgan fingerprints + conditions)", "Conditions + metal (Track A); conditions + ECFP + ligand (Track B)"],
    ["Test design", "One fixed 494-row molecule-held-out test", "Cross-validation over all rows (molecule-grouped for Track A)"],
    ["3-class accuracy, all rows", "0.72 (macro-F1 0.67)", "0.625 new molecule (Track A); 0.753 known molecule (Track B)"],
    ["3-class accuracy, top 25% confidence", "Not available, no ranking", "0.854 (Track A); 0.956 (Track B)"],
    ["3-class accuracy, top 10% confidence", "Not available, no ranking", "0.912 (Track A); 0.983 (Track B)"],
    ["Binary extract screen (logD > 0)", "Not reported", "Track A accuracy 0.742, ROC AUC 0.817; Track B accuracy 0.832, ROC AUC 0.910"],
    ["Continuous logD, R2 and RMSE", "Not provided (classifier only)", "Track A 0.466 and 1.148; Track B 0.725 and 0.823"],
    ["Per-prediction uncertainty", "None", "Confidence score plus conformal interval at 90 percent coverage"],
    ["His XGBoost under our cross-validation", "0.72 reported (one 494-row holdout)", "0.605 over all new molecules; same model on honest CV, so 0.72 reflects the favorable holdout and the imbalanced classes, not a stronger model"],
    ["His XGBoost plus our confidence", "0.72 flat, no ranking", "0.734 at top 50 percent; 0.847 at top 25 percent; 0.914 at top 10 percent; binary screen ROC AUC 0.798"],
    ["Our model on his exact 494-row test", "his 0.72 (tuned XGBoost on this test)", "0.692 raw, which matches him; with our confidence 0.863 at top 25 percent and 1.000 at top 10 percent; also returns continuous logD (R2 0.360) and an interval"],
    ["Common-split test (same data and features)", "his XGBoost: 0.648 our CV, 0.680 his holdout (0.72 tuned)", "our LightGBM: 0.657 our CV, 0.692 his holdout; the two models are about equal on a common split, so the split not the model drives his headline"],
    ["Verdict", "Higher raw 3-class accuracy on its single holdout, but flat, with no way to rank predictions", "Matches on new molecules once confidence is used (0.912 at top 10 percent), beats 0.72 outright on known molecules (0.753), and also returns the value, the uncertainty, and the separations"],
    ],
    [24, 46, 58],
    note="Both datasets have exactly 8075 rows, so they are almost certainly the same integrated f-element dataset. His 3-class cut points are inferred as distribution coefficient D = 0.5 and 10 from the folder name and num_class=3. His accuracy is one 494-row holdout; ours is cross-validated. Therefore this is a close comparison, not an exact one.")

# 8. Methods and models
sheet("Methods and Models",
    ["Item", "Detail"],
    [
    ["Target", "logD, the base-10 log of the distribution coefficient. Range about 17.5 log units."],
    ["Two tasks", "Track A screens new molecules (molecule-grouped CV). Track B optimizes conditions for known molecules (random-row CV). Kept separate so molecule leakage does not inflate the number."],
    ["Models tried", "LightGBM, XGBoost, CatBoost, HistGradientBoosting, ExtraTrees, ridge, small MLP, Chemprop D-MPNN graph network, TabPFN. Final: single LightGBM (A) and an NNLS stack of trees (B)."],
    ["Why a stack", "A non-negative least squares stack weights members and zeros the ones that do not help. It beats the best single model on accuracy. On Track A it costs confidence ranking, so the single model is used there."],
    ["Confidence", "A second model (err_lgb) predicts the absolute error of each prediction from the conditions and the prediction. Rows are ranked by predicted error. The top 10 to 25 percent by confidence are much more accurate (Track B top 10% R2 0.940, RMSE 0.341)."],
    ["Uncertainty intervals", "Normalized split-conformal. Interval width scales with the confidence error and is calibrated to hit 90 and 80 percent coverage. Measured coverage matches the target."],
    ["Data cleaning", "Dropped exact duplicate rows and replicate groups (same molecule, metal, conditions) that disagree by more than 2 log units. 8074 rows to 7065."],
    ["Ceiling", "Within-replicate scatter sets a noise floor near RMSE 0.77 on all data and lower on the cleaned set. Track A is near its achievable ceiling for new molecules given only about 300 distinct extractants."],
    ["TabPFN status", "Tested as local experts inside the stack. Result row T2 in the experiment log. Included only if it earns NNLS weight."],
    ],
    [22, 96])

import os
if os.path.exists("classifier_confidence_results.csv"):
    cc = pd.read_csv("classifier_confidence_results.csv")
    sheet("Classifier with Confidence",
        ["Track", "Task", "Accuracy", "Macro-F1 or AUC", "Top 25% conf accuracy", "Top 10% conf accuracy"],
        cc.values.tolist(), [30, 16, 10, 16, 20, 20],
        note="Native classifiers at the Zhang-style cut points (distribution coefficient 0.5 and 10) and a binary logD greater than 0 screen, each with a confidence score (the max class probability). Selective accuracy rises with confidence, which a single flat accuracy does not show. Zhang reference: 3-class accuracy 0.72, macro-F1 0.67.")

wb.save("REE_Results_Organized.xlsx")
print("saved REE_Results_Organized.xlsx with sheets:", wb.sheetnames)
