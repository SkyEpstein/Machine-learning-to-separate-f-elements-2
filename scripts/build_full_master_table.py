#!/usr/bin/env python3
"""
build_full_master_table.py — assemble the comprehensive master results table from the
extracted rows (master_rows.json). Emits master_results_table.csv (every result), a
grouped markdown RESULTS_TABLE.md, and a "Master results (full)" sheet in the workbook.
Each row carries the result, our model's confidence-filtered companion, the regime, a
caveat, the reasoning behind the decision, and the source. Rows keep their fine-grained
area; a coarse Theme is added for grouping and navigation.
"""
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
SCR = "/private/tmp/claude-501/-Users-skyepstein-Downloads-ML-Model-Folder-8/6289a884-4b42-4d19-b509-c744b27fe645/scratchpad"
rows = json.load(open(f"{SCR}/master_rows.json"))
THEMES = ["logD tracks & deployment", "Free energy (delta G)", "Confidence, classifiers & per-metal",
          "Separation factor", "Active analysis & screening", "Zhang comparison"]
def theme(a):
    s = (a or "").lower()
    if "zhang" in s: return "Zhang comparison"
    if "separation" in s or "sep_factor" in s or "sep factor" in s or "sign correct" in s: return "Separation factor"
    if "dg" in s or "delta g" in s or "free energy" in s: return "Free energy (delta G)"
    if "active" in s or "ucb" in s or "triage" in s or "picks" in s: return "Active analysis & screening"
    if "classifier" in s or "per-metal" in s or "per metal" in s or "confidence recipe" in s: return "Confidence, classifiers & per-metal"
    return "logD tracks & deployment"
for r in rows: r["theme"] = theme(r.get("area", ""))
COLS = [("theme", "Theme"), ("area", "Area"), ("experiment", "Experiment"), ("model", "Model"),
        ("metric", "Metric"), ("value", "Value"), ("rmse", "RMSE"), ("n", "n"), ("regime", "Regime"),
        ("confidence_companion", "Our-model confidence"), ("caveat", "Caveat"),
        ("reasoning", "Reasoning (why this decision)"), ("source", "Source")]
keys = [k for k, _ in COLS]; heads = [h for _, h in COLS]
# stable sort by theme order then keep extraction order within
rows.sort(key=lambda r: THEMES.index(r["theme"]) if r["theme"] in THEMES else 99)
pd.DataFrame([{h: (r.get(k, "") or "") for k, h in COLS} for r in rows]).to_csv("master_results_table.csv", index=False)
# grouped markdown by coarse theme
md = ["# Master results table (comprehensive)", "",
      f"Every result we produced ({len(rows)} rows across {len(set(r['theme'] for r in rows))} themes), each with its metric, our model's confidence-filtered companion where one exists, the evaluation regime, a caveat, the reasoning behind the decision, and the source. All numbers are the audited honest values; the irreducible label-noise floor is about 0.45 log units / 6 kJ/mol. Per project rule, our model's confidence-filtered figure is shown alongside every applicable result, since the calibrated confidence layer is the differentiator.", ""]
disp = [h for h in heads if h != "Theme"]
for th in THEMES:
    sub = [r for r in rows if r["theme"] == th]
    if not sub: continue
    md.append(f"## {th}  ({len(sub)})")
    md.append("| " + " | ".join(disp) + " |")
    md.append("|" + "|".join(["---"] * len(disp)) + "|")
    for r in sub:
        cells = [str(r.get(k, "") or "").replace("|", "\\|").replace("\n", " ") for k, h in COLS if h != "Theme"]
        md.append("| " + " | ".join(cells) + " |")
    md.append("")
open("RESULTS_TABLE.md", "w").write("\n".join(md) + "\n")
# workbook sheet
wbp = "REE_Results_Organized.xlsx"
wb = openpyxl.load_workbook(wbp)
name = "Master results (full)"
if name in wb.sheetnames: del wb[name]
ws = wb.create_sheet(name, 1)
hb = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="2C7FB8")
for j, h in enumerate(heads, 1):
    c = ws.cell(1, j, h); c.font = hb; c.fill = fill; c.alignment = Alignment(wrap_text=True, vertical="top")
for i, r in enumerate(rows, 2):
    for j, (k, h) in enumerate(COLS, 1):
        ws.cell(i, j, str(r.get(k, "") or "")).alignment = Alignment(wrap_text=True, vertical="top")
widths = {"Theme": 22, "Area": 30, "Experiment": 30, "Model": 18, "Metric": 13, "Value": 11, "RMSE": 9, "n": 8, "Regime": 24, "Our-model confidence": 28, "Caveat": 40, "Reasoning (why this decision)": 70, "Source": 26}
for j, h in enumerate(heads, 1): ws.column_dimensions[ws.cell(1, j).column_letter].width = widths.get(h, 18)
ws.freeze_panes = "A2"
wb.save(wbp)
print(f"wrote master_results_table.csv and RESULTS_TABLE.md ({len(rows)} rows) and workbook sheet '{name}'")
print("by theme:", {th: sum(1 for r in rows if r['theme'] == th) for th in THEMES})
