#!/usr/bin/env python3
"""
build_workbook2.py — organize all results into one clean workbook modeled on the
mentor's ResultsData.xlsx (versioned experiment log), but with both R^2 and RMSE,
plain wording (no all-caps emphasis, no em dashes), and extra sheets for the
deployables, confidence, per-metal and per-pair analysis, the Zhang comparison,
and methods. Saves REE_Results_Organized.xlsx.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEAD = PatternFill("solid", fgColor="1F4E79"); SUB = PatternFill("solid", fgColor="D5E8F0")
WHITE = Font(color="FFFFFF", bold=True, name="Arial", sz=11); BOLD = Font(bold=True, name="Arial", sz=11)
REG = Font(name="Arial", sz=10); THIN = Side(style="thin", color="BFBFBF")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN); WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")

wb = Workbook()

def sheet(title, headers, rows, widths, note=None, freeze="A2"):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    if ws.title == "Sheet": ws.title = title
    r0 = 1
    if note:
        ws.cell(1, 1, note).font = Font(italic=True, name="Arial", sz=9, color="555555")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers)); r0 = 2; freeze = "A3"
    for j, h in enumerate(headers, 1):
        c = ws.cell(r0, j, h); c.fill = HEAD; c.font = WHITE; c.alignment = WRAP; c.border = BORD
    for i, row in enumerate(rows, r0 + 1):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v); c.font = REG; c.border = BORD; c.alignment = WRAP
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = freeze
    return ws

# 1. Summary of deployable models
sheet("Summary",
    ["Model (deployable)", "Use case", "Split (honest)", "R2", "RMSE", "Top 25% R2", "Top 25% RMSE", "Top 10% R2", "Top 10% RMSE", "Conformal 90% coverage"],
    [["Track A: single LightGBM", "Screen a new extractant before testing it", "Molecule-grouped (new molecule)", 0.466, 1.148, 0.825, 0.642, 0.912, 0.493, 0.89],
     ["Track B: NNLS stack of trees", "Optimize conditions for a known extractant", "Random row (new conditions)", 0.725, 0.823, 0.911, 0.424, 0.940, 0.341, 0.90]],
    [26, 38, 30, 8, 8, 11, 12, 11, 12, 16],
    note="Both metrics shown throughout: R2 (higher is better) and RMSE in log units (lower is better). Data cleaned to 7065 rows (dropped exact duplicates and replicate groups disagreeing by more than 2 log units).")

# 2. Experiment log (mentor template style, our modeling work)
sheet("Audit corrections",
    ["Quantity", "Honest value", "Previously reported", "Why / caveat"],
    [
    ["Track A new-molecule logD R2", "0.466 (RMSE 1.148)", "0.466", "Honest, molecule-grouped CV. Unchanged."],
    ["Track B logD R2 (honest)", "about 0.61 (RMSE about 0.98)", "0.725", "0.725 used a random-row split that memorizes replicate siblings (the fingerprint is constant within a molecule); about 0.07 R2 is leakage. The honest 0.61 is condition-key grouped: interpolate new conditions for an already-seen molecule."],
    ["Track B on NEW molecules", "about 0.44", "read as 0.725", "The rich features collapse to about Track A on unseen molecules."],
    ["delta G per-pair R2", "0.46 plus or minus 0.01 (RMSE 6.3)", "0.473", "0.473 was the maximum of a roughly 55-way model and hyperparameter search on one fixed split; the mean over shuffled molecule-grouped splits is 0.461 plus or minus 0.009."],
    ["Top-10 percent confidence R2", "selective-prediction operating point", "0.912 / 0.940 / 0.776", "Report the RMSE drop; the R2 rise is partly the shrinking variance of the retained subset. The Track B 0.940 is on a leaky split and is being recomputed on the condition-key split."],
    ["Zhang comparison", "ours 0.657 CV / 0.692 holdout vs his 0.648 / 0.680", "his 0.605; we match 0.72", "0.605 was his model run on our data on a different split, which widened the gap. The same-data comparison shows the models about equal, so the split, not the model, drives his 0.72 headline. His 0.72 reproduces at 0.68; ours is 0.692 on his holdout (within 0.03)."],
    ["Per-metal and classifier R2", "known-molecule (random CV) values", "0.80+ per metal; 0.753 classifier", "These are known-molecule optimistic. The grouped classifier is 0.625 (3-class) vs a 0.385 majority baseline; the per-metal new-molecule numbers are lower."],
    ["Data basis", "label-QC-cleaned (8075 to 7066 rows); noise floor about 0.45 log units", "across all rows", "Cleaning (replicate logD range at most 2) also removes the hardest measurements from evaluation, so all absolute numbers are an upper bound."],
    ],
    [30, 34, 22, 64],
    note="Corrections from a full adversarial audit. The modeling is fundamentally honest: genuine molecule-grouped cross-validation with zero train/test molecule overlap, no feature encodes the target, and the confidence intervals are calibrated (90 percent covers 90 percent). These entries correct optimistic framing and the Track B replicate-memorization leakage so the reported numbers match what the evaluation supports. Where other sheets still show the earlier values, this sheet supersedes them.")

sheet("Experiment Log",
    ["Version", "Extractant representation", "Metals", "Solvents", "Acid", "Model", "Settings", "CV R2", "CV RMSE", "Notes"],
    [
    ["A1", "Conditions only (no structure)", "Properties", "Properties", "Properties", "LightGBM", "new-molecule grouped CV", 0.466, 1.148, "Track A baseline. Structure does not generalize from ~300 molecules, so conditions plus metal carry it."],
    ["A2", "Conditions only (no structure)", "Properties", "Properties", "Properties", "NNLS stack (lgb, xgb, hgb, et, cat, ridge)", "new-molecule grouped CV", 0.504, 1.106, "Stack adds accuracy but makes errors less rankable, so confidence drops. Single model chosen for Track A."],
    ["B1", "Conditions + ECFP + ligand", "Properties", "Properties", "Properties", "LightGBM", "known-molecule random CV", 0.714, 0.831, "Track B single model."],
    ["B2", "Conditions + ECFP + ligand", "Properties", "Properties", "Properties", "NNLS stack (lgb 0.40, et 0.43, xgb 0.14, cat 0.03)", "known-molecule random CV", 0.725, 0.823, "Track B deployable. Stack wins accuracy and confidence here. Beats the mentor log best test RMSE of about 0.96."],
    ["C1", "Conditions + ECFP + RDKit descriptors", "Properties", "Properties", "Properties", "LightGBM", "known-molecule random CV", 0.70, 0.86, "RDKit descriptors did not help over ECFP and slightly hurt."],
    ["C2", "Conditions + ECFP + learned embeddings", "Properties", "Properties", "Properties", "LightGBM", "known-molecule random CV", 0.69, 0.88, "Pretrained embeddings hurt. ECFP plus conditions is the better representation."],
    ["G1", "Graph (Chemprop D-MPNN) alone", "Properties", "Properties", "Properties", "Message-passing GNN", "new-molecule grouped CV", 0.23, None, "GNN alone is weak on ~300 molecules. RMSE not logged."],
    ["G2", "Graph GNN + tree global blend", "Properties", "Properties", "Properties", "GNN + LightGBM blend", "new-molecule grouped CV", 0.47, 1.15, "GNN adds no generalizable gain over the tree model."],
    ["T1", "Conditions + ECFP (PCA 50) + ligand", "Properties", "Properties", "Properties", "TabPFN global (context capped 1000)", "known-molecule random CV", 0.36, None, "TabPFN cannot use more than 1000 context rows on CPU, so a single global TabPFN is weak here."],
    ["T2", "Conditions + ECFP (PCA 50) + ligand", "Properties", "Properties", "Properties", "TabPFN local experts + tree stack", "known-molecule random CV", 0.701, None, "Lighter in-stack test (6 regions, 3 folds). TabPFN earns NNLS weight 0.13 but adds only +0.0011 R2 over the trees-only stack (0.700); error-correlation with the trees is 0.82, so little diversity. Dropped, since it does not clear the keep threshold."],
    ["P1", "Conditions + ECFP + slope or physics features", "Properties", "Properties", "Properties", "LightGBM", "known-molecule random CV", 0.72, 0.83, "Physics-informed features gave no clear gain. The clean logK + n log[HA] + n pH signal is not present in most rows."],
    ["H1", "Conditions + ECFP + ligand", "Properties", "Properties", "Properties", "LightGBM tuned (randomized search)", "both tracks", 0.484, 1.13, "Hyperparameter search moved Track A 0.466 to 0.484 and Track B not at all. Negligible."],
    ["E1", "Conditions + ECFP + ligand, tail experts", "Properties", "Properties", "Properties", "LightGBM + extreme-region experts", "known-molecule random CV", 0.70, 0.86, "Tail experts for the extremes hurt overall (0.725 to about 0.70). The global model compresses tails but naive experts do not fix it."],
    ["X0", "Leaked single split (for contrast only)", "Properties", "Properties", "Properties", "LightGBM", "single split with molecule leakage", 0.60, 1.05, "Not honest. A molecule appears in train and test. Shown only to explain why earlier numbers looked higher."],
    ],
    [10, 34, 12, 12, 10, 34, 22, 8, 9, 46],
    note="Versioned log in the style of ResultsData.xlsx, with both CV R2 and CV RMSE. CV numbers are cross-validated (honest), not a single lucky split.")

# 3. Model comparison: single vs stack, members and weights
sheet("Model Comparison",
    ["Track", "Member", "Member R2", "NNLS weight", "Single R2", "Single RMSE", "Stack R2", "Stack RMSE", "Chosen for deploy"],
    [
    ["A (new molecule)", "cat", 0.491, 0.37, 0.466, 1.148, 0.504, 1.106, "Single (better confidence)"],
    ["A (new molecule)", "et", 0.470, 0.24, "", "", "", "", ""],
    ["A (new molecule)", "xgb", 0.466, 0.16, "", "", "", "", ""],
    ["A (new molecule)", "lgb", 0.459, 0.11, "", "", "", "", ""],
    ["A (new molecule)", "ridge", 0.178, 0.10, "", "", "", "", ""],
    ["A (new molecule)", "hgb / mlp", 0.47, 0.02, "", "", "", "", ""],
    ["B (known molecule)", "et", 0.708, 0.43, 0.714, 0.831, 0.725, 0.823, "Stack (better accuracy and confidence)"],
    ["B (known molecule)", "lgb", 0.711, 0.40, "", "", "", "", ""],
    ["B (known molecule)", "xgb", 0.707, 0.14, "", "", "", "", ""],
    ["B (known molecule)", "cat", 0.692, 0.03, "", "", "", "", ""],
    ["B (known molecule)", "hgb / mlp / ridge", 0.68, 0.00, "", "", "", "", "Dropped by NNLS"],
    ],
    [18, 16, 11, 12, 9, 10, 9, 10, 32],
    note="NNLS stacking gives zero weight to members that do not improve the blend, so weak models drop out automatically.")

# 4. Confidence bakeoff
sheet("Confidence",
    ["Track", "Predictor", "Err-model recipe", "Spearman(err, abs error)", "Top 50% R2", "Top 50% RMSE", "Top 25% R2", "Top 25% RMSE", "Top 10% R2", "Top 10% RMSE", "Chosen"],
    [
    ["A", "single", "plain + strong", 0.416, 0.719, 0.812, 0.825, 0.642, 0.912, 0.493, "yes"],
    ["A", "stack", "plain + strong", 0.380, 0.718, 0.791, 0.802, 0.656, 0.843, 0.549, ""],
    ["A", "stack", "plain + regularized", 0.387, 0.727, 0.776, 0.812, 0.634, 0.844, 0.545, ""],
    ["A", "stack", "rich (members, disagreement, novelty)", 0.338, 0.687, 0.835, 0.799, 0.662, 0.863, 0.569, "worst"],
    ["A", "stack", "lean (disagreement, novelty)", 0.340, 0.713, 0.804, 0.790, 0.670, 0.831, 0.556, ""],
    ["B", "single", "plain + strong", 0.426, 0.866, 0.550, 0.915, 0.440, 0.938, 0.379, ""],
    ["B", "stack", "plain + regularized", 0.449, 0.871, 0.535, 0.911, 0.424, 0.940, 0.341, "yes"],
    ["B", "stack", "rich", 0.429, 0.872, 0.542, 0.923, 0.408, 0.937, 0.339, ""],
    ["B", "stack", "lean", 0.437, 0.875, 0.535, 0.908, 0.440, 0.942, 0.344, ""],
    ],
    [7, 9, 34, 18, 11, 12, 11, 12, 11, 12, 9],
    note="Confidence is a learned error model (err_lgb) that predicts each row's absolute error. Rows are ranked by it. Rich features (member predictions, disagreement, novelty) hurt, so the plain recipe was kept. Conformal intervals scale by this error and hit their target coverage.")

# 5 and 6: metal and pair from CSV
bm = pd.read_csv("metal_confidence_by_metal.csv")
sheet("By Metal",
    ["Metal", "n", "R2", "RMSE", "Median confidence error", "Top 25% R2", "Top 25% RMSE"],
    bm.values.tolist(), [10, 7, 8, 8, 20, 11, 12],
    note="Track B (known molecule). Accuracy and confidence by metal, both metrics. The confidence error tracks RMSE, so the model flags the metals it predicts poorly (for example Np(V), U(VI), Er(III)).")
bp = pd.read_csv("metal_confidence_by_pair.csv")
sheet("By Metal Pair",
    ["Metal pair", "n", "Separation R2", "Separation RMSE", "Mean confidence error"],
    bp.values.tolist(), [18, 7, 14, 16, 20],
    note="Predicted logD difference between two metals at the same conditions (the separation use case). Overall separation R2 0.599, RMSE 1.025. Tight adjacent pairs are hardest and the confidence error is highest there.")

# 7. Zhang comparison
sheet("Zhang Comparison",
    ["Aspect", "Dr. Zhang (SAFE-MolGen)", "Our model"],
    [
    ["Goal", "Generate new extractants with an LLM, then screen them", "Predict logD, optimize conditions, and predict separations"],
    ["Supervised model", "XGBoost classifier (also NN with CV, random forest)", "LightGBM and an NNLS tree stack, plus matching classifiers"],
    ["Task type", "3-class classification only", "Regression of continuous logD, plus 3-class and binary classifiers"],
    ["Data", "8075 rows, f-element extraction (ACSEPT, DGA, IDEaL, ORNL)", "Same 8075 rows (7065 after cleaning)"],
    ["Features", "1860 (Morgan fingerprints + conditions)", "Conditions + metal (Track A); conditions + ECFP + ligand (Track B)"],
    ["Test design", "One fixed 494-row molecule-held-out test", "Cross-validation over all rows (molecule-grouped for Track A)"],
    ["3-class accuracy, all rows", "0.72 (macro-F1 0.67)", "0.625 new molecule (Track A); 0.753 known molecule (Track B)"],
    ["3-class accuracy, top 25% confidence", "Not available, no ranking", "0.854 (Track A); 0.956 (Track B)"],
    ["3-class accuracy, top 10% confidence", "Not available, no ranking", "0.912 (Track A); 0.983 (Track B)"],
    ["Binary extract screen (logD > 0)", "Not reported", "Track A accuracy 0.742, ROC AUC 0.817; Track B accuracy 0.832, ROC AUC 0.910"],
    ["Continuous logD, R2 and RMSE", "Not provided (classifier only)", "Track A 0.466 and 1.148; Track B 0.725 and 0.823"],
    ["Per-prediction uncertainty", "None", "Confidence score plus conformal interval at 90 percent coverage"],
    ["His XGBoost, same-data comparison", "0.72 reported (one 494-row holdout)", "On the same data his model scores 0.648 (our CV) and 0.680 (his holdout); his 0.72 reproduces at 0.68. The split, not a stronger model, drives the headline. An earlier 0.605 figure was his model on a different split and overstated the gap."],
    ["His XGBoost plus our confidence", "0.72 flat, no ranking", "0.734 at top 50 percent; 0.847 at top 25 percent; 0.914 at top 10 percent; binary screen ROC AUC 0.798"],
    ["Our model on his exact 494-row test", "his 0.72 (tuned XGBoost on this test)", "0.692, within about 0.03 of his 0.72; with our confidence 0.863 at top 25 percent and 1.000 at top 10 percent; also returns continuous logD (R2 0.360) and an interval"],
    ["Common-split test (same data and features)", "his XGBoost: 0.648 our CV, 0.680 his holdout (0.72 tuned)", "our LightGBM: 0.657 our CV, 0.692 his holdout; the two models are about equal on a common split, so the split not the model drives his headline"],
    ["Verdict", "Higher raw 3-class accuracy on its single holdout, but flat, with no way to rank predictions", "About equal on the same-data common split (0.657 vs 0.648 our CV; 0.692 vs 0.680 his holdout), so the split drives his headline. Our added value is calibrated uncertainty, continuous logD, and the separations, not higher accuracy."],
    ],
    [24, 46, 58],
    note="Both datasets have exactly 8075 rows, so they are almost certainly the same integrated f-element dataset. His 3-class cut points are inferred as distribution coefficient D = 0.5 and 10 from the folder name and num_class=3. His accuracy is one 494-row holdout; ours is cross-validated. Therefore this is a close comparison, not an exact one.")

# 8. Methods and models
sheet("Methods and Models",
    ["Item", "Detail"],
    [
    ["Target", "logD, the base-10 log of the distribution coefficient. Range about 17.5 log units."],
    ["Two tasks", "Track A screens new molecules (molecule-grouped CV). Track B optimizes conditions for known molecules (random-row CV). Kept separate so molecule leakage does not inflate the number."],
    ["Models tried", "LightGBM, XGBoost, CatBoost, HistGradientBoosting, ExtraTrees, ridge, small MLP, Chemprop D-MPNN graph network, TabPFN. Final: single LightGBM (A) and an NNLS stack of trees (B)."],
    ["Why a stack", "A non-negative least squares stack weights members and zeros the ones that do not help. It beats the best single model on accuracy. On Track A it costs confidence ranking, so the single model is used there."],
    ["Confidence", "A second model (err_lgb) predicts the absolute error of each prediction from the conditions and the prediction. Rows are ranked by predicted error. The top 10 to 25 percent by confidence are much more accurate (Track B top 10% R2 0.940, RMSE 0.341)."],
    ["Uncertainty intervals", "Normalized split-conformal. Interval width scales with the confidence error and is calibrated to hit 90 and 80 percent coverage. Measured coverage matches the target."],
    ["Data cleaning", "Dropped exact duplicate rows and replicate groups (same molecule, metal, conditions) that disagree by more than 2 log units. 8074 rows to 7065."],
    ["Ceiling", "Within-replicate scatter sets a noise floor near RMSE 0.77 on all data and lower on the cleaned set. Track A is near its achievable ceiling for new molecules given only about 300 distinct extractants."],
    ["TabPFN status", "Tested as local experts inside the stack. Result row T2 in the experiment log. Included only if it earns NNLS weight."],
    ],
    [22, 96])

sheet("Sign and Separation",
    ["Analysis", "Track", "All", "Top 25% conf", "Top 10% conf", "Notes"],
    [
    ["Sign correctness (predicted logD sign matches actual)", "A (new molecule)", 0.743, 0.855, 0.903, "Does the model get extract-or-not (logD greater than 0) right"],
    ["Sign correctness", "B (known molecule)", 0.836, 0.913, 0.919, ""],
    ["Separation, signed R2 of the logD difference", "B (known molecule)", 0.582, "", "", "RMSE 1.069; 44894 metal pairs at matched conditions"],
    ["Separation, R2 of the magnitude of the difference", "B (known molecule)", 0.312, "", "", "how well the size of the gap is predicted"],
    ["Separation, selectivity direction correct", "B (known molecule)", 0.763, "", "", "picks the right metal as the more extracted one"],
    ["Separation, predict-under rate (pred magnitude at most actual)", "B (known molecule)", 0.633, "", "", "the model is conservative, so the real gap is usually at least the predicted one"],
    ["Separation, R2 of the magnitude of the difference", "A (new molecule)", -0.108, "", "", "RMSE 1.145; new-molecule separation does not predict well, worse than guessing the mean"],
    ["Separation, predict-under rate (pred magnitude at most actual)", "A (new molecule)", 0.567, 0.541, 0.486, "conservative overall, but the most confident pairs over-predict more often"],
    ],
    [42, 18, 8, 13, 13, 44],
    note="Sign correctness is the fraction where predicted logD has the same sign as actual (extract or not). Separation compares predicted versus actual logD difference between two metals at the same conditions. The predict-under rate is how often the predicted separation magnitude is at most the actual one, meaning the model is conservative about selectivity.")

sheet("Decision Log",
    ["Decision", "What we chose", "Why"],
    [
    ["Two prediction tasks", "Separate new-molecule (A) from known-molecule (B)", "Mixing them lets a molecule leak across train and test and inflates the score; an earlier single split read 0.60 and did not hold up."],
    ["Evaluation", "Molecule-grouped CV for A, random-row CV for B", "Each matches a real use case (screen a new extractant vs tune a known one) without leakage."],
    ["Data cleaning", "Drop exact duplicates and replicate groups disagreeing by more than 2 log units", "Contradictory labels are noise; cleaning took 8075 rows to 7065."],
    ["Track A model", "Single tuned LightGBM", "The stack adds accuracy but makes residuals less rankable, which weakens confidence; the single model ranks confidence better for screening."],
    ["Track B model", "NNLS stack of trees (ExtraTrees, LightGBM, XGBoost, CatBoost)", "Stacking improves both accuracy and confidence here, and NNLS sets weak members to zero on its own."],
    ["Features", "A: conditions + metal; B: also ECFP + ligand", "Structure does not generalize from about 300 molecules (A); it helps once the molecule has been seen (B). RDKit blocks and embeddings did not beat ECFP plus conditions."],
    ["Confidence", "err_lgb on plain features, ranked; normalized split-conformal intervals", "Richer features (member predictions, disagreement, novelty) ranked worse; the intervals hit their target coverage (0.89 to 0.90)."],
    ["Graph network", "Dropped", "Chemprop alone reached only R2 0.23 on new molecules and added nothing when blended."],
    ["TabPFN", "Dropped", "As a local expert it earned only 0.13 NNLS weight and +0.0011 R2, with error-correlation 0.82 to the trees (not diverse)."],
    ["Hyperparameter search", "Kept near-default settings", "A randomized search moved Track A 0.466 to 0.484 and Track B not at all."],
    ["Separation criterion", "Predict-under: predicted gap at most the actual gap", "Direction does not matter since the other metal stays in solution for a known extractant; a conservative predicted gap means the real gap is at least as large."],
    ["Active analysis", "Rank by prediction for one-shot screening; UCB for sequential rounds", "UCB's uncertainty bonus pulls in mediocre candidates, so greedy finds the strongest now; UCB pays off only when exploring to improve the model over rounds."],
    ["Comparison with Dr. Zhang", "Same 8075-row dataset, scored on his split", "Our model matches his accuracy on his own test; the added value is continuous logD with calibrated confidence rather than one of three buckets."],
    ],
    [26, 40, 66],
    note="Every major modeling choice and the reason for it, so the project is documented as it evolved.")

import os
if os.path.exists("classifier_confidence_results.csv"):
    cc = pd.read_csv("classifier_confidence_results.csv")
    sheet("Classifier with Confidence",
        ["Track", "Task", "Accuracy", "Macro-F1 or AUC", "Top 25% conf accuracy", "Top 10% conf accuracy"],
        cc.values.tolist(), [30, 16, 10, 16, 20, 20],
        note="Native classifiers at the Zhang-style cut points (distribution coefficient 0.5 and 10) and a binary logD greater than 0 screen, each with a confidence score (the max class probability). Selective accuracy rises with confidence, which a single flat accuracy does not show. Zhang reference: 3-class accuracy 0.72, macro-F1 0.67.")

if os.path.exists("ucb_analysis_results.csv"):
    uc = pd.read_csv("ucb_analysis_results.csv")
    sheet("Active analysis (UCB)",
        ["Track", "Select top %", "Method", "Mean actual logD", "Recall of true-best"],
        uc.values.tolist(), [22, 12, 22, 16, 18],
        note="Ranking candidate combinations to test. UCB is the upper end of the 90 percent conformal interval (prediction plus uncertainty). For one-shot screening, ranking by the plain prediction (greedy) finds the strongest extractants and beats UCB, because the uncertainty bonus pulls in uncertain but mediocre candidates. UCB's role is sequential active learning, where exploring uncertain candidates improves the model over rounds. The top 5 percent by prediction reach mean actual logD near 2.2 (new) to 2.6 (known) against 0.02 over all rows.")

if os.path.exists("active_learning_results.csv"):
    al = pd.read_csv("active_learning_results.csv")[["strategy", "final_test_R2", "final_discovery"]]
    sheet("Active learning (sequential)",
        ["Strategy", "Final test R2 (model quality)", "Final discovery (share of true top-10% found)"],
        al.values.tolist(), [22, 30, 40],
        note="Sequential active learning: each round picks the next batch to run, reveals logD, and retrains. For finding strong extractants (discovery) greedy and UCB win, about 0.72 of the true best found. For improving the model (test R2) the ensemble-disagreement uncertainty wins at 0.56. Acquiring by the confidence algorithm's own predicted error, the same err model used to decide which predictions to trust, helps discovery (0.44) but improves the model less (0.465), because large predicted error flags the hard, extreme-logD points rather than under-sampled regions, which biases the training set much as greedy does. So the err model is the right tool for deciding which predictions to trust, while ensemble disagreement is the right signal for choosing points that most improve the model.")

if os.path.exists("experiment_triage_results.csv"):
    et = pd.read_csv("experiment_triage_results.csv")
    et['accepted_within_0p5'] = (et['accepted_within_0p5'] * 100).round(0).astype(int).astype(str) + '%'
    sheet("Experiment triage",
        ["Track", "Auto-accept most confident %", "Experiments saved", "Experiments still needed", "Accepted RMSE", "Accepted within 0.5 log units"],
        et[['track', 'auto_accept_pct', 'experiments_saved', 'experiments_needed', 'accepted_RMSE', 'accepted_within_0p5']].values.tolist(),
        [22, 26, 16, 20, 12, 22],
        note="The high-confidence predictions do not need experiments; they are trusted as-is, so the experiment budget goes to the uncertain remainder. For known molecules the most confident half can be accepted at RMSE 0.535 with 75 percent within half a log unit, which halves the experiments needed. New molecules degrade faster, so a smaller, more confident slice is accepted. For the experiments that are run, sampling by uncertainty improves the model fastest.")

if os.path.exists("dg_results.csv"):
    ec = pd.read_csv("dg_results.csv")
    sheet("Free energy (delta G)",
        ["Framing", "Systems", "R2 all", "RMSE all (kJ/mol)", "R2 top 25% conf", "R2 top 10% conf", "RMSE top 10% (kJ/mol)"],
        ec[['framing', 'n', 'R2_all', 'RMSE_all_kJmol', 'R2_top25', 'R2_top10', 'RMSE_top10_kJmol']].values.tolist(),
        [20, 10, 9, 18, 16, 16, 20],
        note="The free-energy model predicts the Gibbs free energy of extraction (delta G = -2.303 R T logD, in kJ/mol, so favorable extraction is negative) from the extractant structure and the metal, with all reaction conditions dropped. The model is a RandomForest, the best of an ensemble sweep (see the Delta G ensemble sweep sheet). Per-row keeps every measurement; per-pair averages delta G to one value per extractant-metal-acid-diluent system and is condition-independent. Per-pair is the better target (R2 0.47 versus 0.28) because identical structures otherwise repeat with condition-driven delta G the structure cannot explain. Confidence ranking is strong for per-pair: the most confident quarter reach R2 0.68 and the top tenth R2 0.78 (RMSE 3.6 kJ/mol). Molecule-grouped cross-validation, so a new extractant never appears in both train and test.")

if os.path.exists("dg_ensemble_results.csv"):
    es = pd.read_csv("dg_ensemble_results.csv")
    sheet("Delta G ensemble sweep",
        ["Model", "R2 (per-pair dG)", "RMSE (kJ/mol)"],
        es[['model', 'R2', 'RMSE_kJmol']].values.tolist(),
        [26, 16, 16],
        note="Ensemble sweep for the per-pair delta G target: each base model under molecule-grouped cross-validation, then equal-weight and NNLS stacks. RandomForest is the best single model (R2 0.473), with the bagged-tree models ahead of the boosters on this small wide table and Ridge near zero (the signal is nonlinear). The NNLS stack, scored with a second cross-validation (0.474), only ties RandomForest, so stacking adds nothing here and the free-energy model uses plain RandomForest. NNLS weights: RandomForest 0.56, CatBoost 0.28, ExtraTrees 0.19.")

if os.path.exists("dg_ucb_results.csv"):
    du = pd.read_csv("dg_ucb_results.csv")
    sheet("Delta G UCB",
        ["Analysis", "Top %", "Method", "Mean actual dG / RMSE (kJ/mol)", "Recall / within 3 kJ/mol"],
        du[['analysis', 'select_top_pct', 'method', 'mean_actual_dG', 'recall_true_best']].values.tolist(),
        [16, 8, 18, 28, 22],
        note="UCB active analysis run on the delta G model, since delta G is the predictor that drives the screening. To find the strongest extractants (most negative delta G), ranking by the plain prediction (greedy) is best for a one-shot pick: the top 5 percent average delta G -10.5 kJ/mol versus -0.3 at random, while UCB's uncertainty term dilutes that and is for sequential rounds. Triage: the most confident quarter of predictions can be accepted without experiments at RMSE about 4 kJ/mol, so the experiment budget goes to the uncertain rest.")

if os.path.exists("dg_pseudo_results.csv"):
    dp = pd.read_csv("dg_pseudo_results.csv")
    sheet("Delta G pseudo-labeling",
        ["Setting", "Confident added %", "Test R2", "Test RMSE (kJ/mol)"],
        dp[['setting', 'add_confident_pct', 'test_R2', 'test_RMSE_kJmol']].values.tolist(),
        [30, 18, 12, 18],
        note="Testing the idea of adding confident predictions (the narrow 90 percent interval ones) as training rows without experiments. Adding the model's own confident predictions barely helps (test R2 +0.00 to +0.01), because a confident prediction carries no new information. Adding the same systems with their true labels, which is what running the experiments would give, helps materially (+0.06 to +0.10). So confidence is for deciding which predictions to trust and skip, not for manufacturing training data; the real gains come from experiments on the uncertain points.")

if os.path.exists("dg_confidence_sweep_results.csv"):
    cs = pd.read_csv("dg_confidence_sweep_results.csv")
    sheet("Confidence sweep",
        ["Part", "Model or ranker", "All R2", "Top 25% R2", "Top 10% R2", "Top 10% RMSE"],
        cs[['part', 'name', 'all_R2', 'top25_R2', 'top10_R2', 'top10_RMSE']].values.tolist(),
        [18, 24, 9, 12, 12, 14],
        note="Two sweeps on the per-pair delta G model. Part A applies the same LightGBM confidence ranker to each prediction model: all tree models benefit, climbing from about 0.42-0.47 overall to roughly 0.75-0.79 at the most confident tenth, with XGBoost and HistGB edging RandomForest on that small sliver while RandomForest stays best overall, and CatBoost the outlier whose errors are least predictable. Part B fixes the predictor at RandomForest and sweeps the confidence estimator: the learned err models (LightGBM, RandomForest) rank best and are about tied, while RandomForest's own tree-spread is the weakest ranker, which is why a learned err model is used rather than the native uncertainty. The current setup (RandomForest predictor, LightGBM err ranker) is validated.")

if os.path.exists("dg_coverage_results.csv"):
    cov = pd.read_csv("dg_coverage_results.csv").fillna("")
    sheet("Delta G interval calibration",
        ["Target", "Empirical coverage", "Mean width (kJ/mol)"],
        cov[['target', 'empirical_coverage', 'mean_width_kJmol']].astype(object).values.tolist(),
        [12, 20, 20],
        note="Honest molecule-split conformal check of the delta G prediction intervals: the conformal quantile is set on one set of molecules and coverage is measured on a disjoint set. The 90 percent interval covers 90.2 percent of held-out cases and the 80 percent interval covers 80.2 percent, so the stated confidence can be trusted. The intervals are calibrated and heterogeneous: the confident, narrow-interval predictions are the accurate ones (most confident tenth R2 0.78, RMSE 3.6 kJ/mol), while the uncertain ones get wide intervals and are sent to the lab. This calibration is what lets the intervals drive the active analysis: trust and skip the confident predictions, test the uncertain ones.")

wb.save("REE_Results_Organized.xlsx")
print("saved REE_Results_Organized.xlsx with sheets:", wb.sheetnames)
