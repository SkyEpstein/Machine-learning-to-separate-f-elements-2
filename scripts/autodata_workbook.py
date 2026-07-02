#!/usr/bin/env python3
"""Append an AutoData / new-extractant-confidence sheet to the workbook, append-only
(load existing, add one sheet, save), never rebuilding the other sheets."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
WB = "docs/REE_Results_Organized.xlsx"
wb = openpyxl.load_workbook(WB)
name = "AutoData & new-ext confidence"
if name in wb.sheetnames: del wb[name]
ws = wb.create_sheet(name)
hdr = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="2E5A88"); wrap = Alignment(wrap_text=True, vertical="top")
cols = ["Result", "Metric", "Value", "Regime", "Caveat", "Source"]
ws.append(cols)
for c in range(1, len(cols) + 1):
    ws.cell(1, c).font = hdr; ws.cell(1, c).fill = fill
rows = [
    ["AutoData generator (Am/Eu): novel vs in-sample known", "pred_sep p90 (separation factor)", "novel 0.99 (~10x) vs known 1.32 (~22x)", "new extractant, de novo", "novel does NOT beat best known; magnitude low-trust", "autodata_pool.py"],
    ["AutoData pool quality (Am/Eu)", "parse-valid / novel / scaffolds", "95% / 105 / 56", "generate-score-refine loop", "novel = new-to-dataset, not new-to-chemistry (Cyanex-301 family)", "autodata_pool.py"],
    ["Selectivity vs strength (known)", "Spearman(pred_sep, general strength)", "-0.20", "known, predicted", "genuine selectivity; S-donors most Am-selective (1.07 vs 0.49)", "sep_selectivity_check.py"],
    ["Selectivity vs strength (novel)", "Spearman(pred_sep, general strength)", "+0.53", "new extractant, predicted", "partly collapses into strength; 99% favor Am", "sep_selectivity_check.py"],
    ["All-pair selectivity structure", "PC1 variance / extractant-specific", "66% / 34%", "known, 28 f-metals", "pairwise selectivity genuine but ~1/3 extractant-specific", "sep_allpairs.py"],
    ["Best-pair TRUSTED shortlist (top-10% conf)", "predicted separation factor", "14 to 32", "new extractant, support>=20 An/Ln pairs", "predicted; needs lab validation; Am/La, Am/Gd, La/Nd", "autodata_bp_pool.py"],
    ["New-ext confidence bake-off WINNER: ens_bag", "top-10% direction acc", "0.66 -> 0.81", "new extractant (molecule-grouped), top-10%", "only method to improve; RMSE fall partly shrinkage", "newext_confidence_bakeoff.py"],
    ["New-ext confidence WINNER: ens_bag", "top-10% signed R2 / useful-F1 / calibration", "0.28 / 0.59 / 0.31", "new extractant, top-10%", "incumbent learned-error: signed R2 -0.00, calibration 0.20", "newext_confidence_bakeoff.py"],
    ["New-ext confidence LOSERS: AD distances", "calibration rho (Tanimoto / desc-kNN)", "0.008 / 0.022", "new extractant", "expected to win but lost; bake-off overturned intuition", "newext_confidence_bakeoff.py"],
    ["Known-ext confidence (unchanged, works)", "top-10% signed R2", "0.36 -> 0.59", "known extractant, top-10%", "the known-extractant layer already works; not modified", "sep_factor_confidence.py"],
]
for r in rows:
    ws.append(r)
    for c in range(1, len(cols) + 1): ws.cell(ws.max_row, c).alignment = wrap
widths = [42, 30, 26, 30, 44, 26]
for i, w in enumerate(widths, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
wb.save(WB)
print(f"added sheet '{name}' ({len(rows)} rows); workbook now has {len(wb.sheetnames)} sheets")
